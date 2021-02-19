import openpyxl as xl

# load the workbook
wb = xl.load_workbook("./ExcelHandling/DSTeamData.xlsx")

# get the sheets
sheets = wb.sheetnames
print(sheets)

# get specific sheet
sht = wb['Members']

# access excel data - method 1
print(sht['A2'].value)

# access excel data - method 2
print(sht.cell(2,2).value)

# access excel data - method 3
print(sht.cell(row = 2, column = 3).value)