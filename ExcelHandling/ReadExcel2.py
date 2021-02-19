import openpyxl as xl

# load the workbook
wb = xl.load_workbook("./ExcelHandling/DSTeamData.xlsx")

# get specific sheet
sht = wb['Members']

# get row and column count
rows = sht.max_row
cols = sht.max_column

# access all data
for i in range(1, rows+1):
    for j in range(1, cols+1):
        if(j != cols):
            print((sht.cell(i, j).value), ",", end="")
        else:
            print((sht.cell(i, j).value), end="")
    print()